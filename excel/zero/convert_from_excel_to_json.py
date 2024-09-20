# МогучийДД (MoguchiyDD)
# 2024.07.28, 12:06:35 PM
# convert_from_excel_to_json.py


from enum import Enum

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter, coordinate_to_tuple

from json import load, dump
from time import sleep

from os import path, remove


class ConvertFromExcelToJSON:
    """
    Convert from EXCEL to JSON

    PARAMETERS:
    - **filepath** : an EXCEL file
    - **json_schema** : a JSON SCHEMA file
    - **result_filename_for_json** : a file name for final JSON file

    FUNCTIONS:
    - **cmd**(type: Commands, sheet: int = 0) -> Workbook | Worksheet |
    list[dict] | None : Running a class via available commands
    """

    class Commands(Enum):
        open = "open"
        sheet = "sheet"
        schema = "schema"
        getData = "getData"
        load = "load"
        getDataAndLoad = "getDataAndLoad"
        close = "close"

    def __init__(
        self,
        filepath: str,
        json_schema: str,
        result_filename_for_json: str = "result-excel"
    ) -> None:
        self.filepath: str = filepath
        self.schema: str = json_schema
        self.result_filename_for_json: str = result_filename_for_json
        self.data_schema: dict = {}
        self.data_for_save: list = []
        self.workbook: None | Workbook = None
        self.worksheet: None | Worksheet = None
        self.sheet: int = 0

    def cmd(
        self,
        type: Commands,
        sheet: int = 0
    ) -> Workbook | Worksheet | list[dict] | None:
        """
        Running a class via available commands

        PARAMETERS:
        - **type** : Variable from the **Command** class
        - **sheet** : Opening sheet number from an EXCEL file

        COMMANDS:
        - «**open**» : Opening an EXCEL file
        - «**sheet**» : Taking a sheet from an EXCEL file for processing
        (default 1st)
        - «**schema**» : Opening a SCHEMA (JSON file)
        - «**getData**» : Get data from an EXCEL file
        - «**load**» : Write data via schema to a new JSON file
        - «**getDataAndLoad**» : Running the «getData» and «load» commands
        - «**close**» : Closing an EXCEL file

        RETURN: the commands «**open**», «**sheet**» and «**getData**» have
        data return, and the rest nothing
        """

        if type is self.Commands.open:
            self.__enter__()
            return self.workbook
        elif type is self.Commands.sheet:
            self.sheet = sheet
            self.__sheet__()
            return self.worksheet
        elif type is self.Commands.schema:
            self.__schema__()
        elif type is self.Commands.getData:
            self.__get_data__()
            return self.data_for_save
        elif type is self.Commands.load:
            self.__save_from_load__()
        elif type is self.Commands.getDataAndLoad:
            self.__get_data__()
            self.__save_from_load__()
        elif type is self.Commands.close:
            self.__exit__()
        
        return None

    def __enter__(self) -> None:
        """
        Opening an EXCEL file \\
        **Launch via the «open» command**
        """

        self.workbook: Workbook = load_workbook(self.filepath, data_only=True)

    def __sheet__(self) -> None:
        """
        Taking a sheet from an EXCEL file for processing (default 1st) \\
        **Launch via the «sheet» command**
        """

        self.worksheet: Worksheet = self.workbook.worksheets[self.sheet]

    def __schema__(self) -> None:
        """
        Opening a SCHEMA (JSON file) \\
        **Launch via the «schema» command**
        """

        with open(self.schema, "r") as schema:
            self.data_schema: dict = load(schema)

    def __get_data__(self) -> None:
        """
        Extracting data from an EXCEL file of a specific 1 sheet into
        a dictionary \\
        **Launch via the «getData» command**
        """

        def find_merge_cells(field: str) -> tuple[int] | None:
            """
            Determine by field whether it is included in the merge of fields

            PARAMETERS:
            - **field** : For example, A9

            RETURN: (minimum row, minimum column, maximum row, maximum column)
            """

            coordinates: tuple[int, int] = coordinate_to_tuple(field)
            row: int = coordinates[0]
            column: int = coordinates[1]

            for merged_cell_range in self.worksheet.merged_cells.ranges:
                min_col, min_row, max_col, max_row = merged_cell_range.bounds
                if min_row <= row <= max_row and min_col <= column <= max_col:
                    merge_cells: tuple[int] = (
                        min_row, min_col,
                        max_row, max_col
                    )
                    return merge_cells
                sleep(0.01)

            return None

        if len(self.data_schema) == 0:
            print("Schema is empty")
            exit(-1)

        # Read Schema
        rows_from: int = self.data_schema["rows_from"]
        rows_to: int = self.data_schema["rows_to"]
        rows_except: list[int] | None = self.data_schema["rows_except"]
        fields_from: str = self.data_schema["fields_from"]
        fields_to: str = self.data_schema["fields_to"]
        fields_except: list[str] | None = [
            self.data_schema["fields_except"]
            if self.data_schema["fields_except"] else []
        ][0]

        # Important Variables
        is_loop: bool = True
        stop_count: int = 1
        count: int = 1
        cur_row: int = rows_from
        fields_data: dict = {}

        while is_loop:  # for 'field_to'
            col_letter: str = get_column_letter(count)
            if col_letter == fields_from:
                is_loop = False
                stop_count, count = count, stop_count
            else:
                count += 1
            sleep(0.01)

        is_loop = True
        while is_loop:  # main
            col_letter = get_column_letter(count)
            if col_letter not in fields_except:  # Save Data from Excel
                field: str = f"{ col_letter }{ cur_row }"
                is_merge_field: tuple[int] | None = find_merge_cells(field)
                if is_merge_field:
                    col_letter = get_column_letter(is_merge_field[1])
                    field = f"{ col_letter }{ is_merge_field[0] }"
                fields_data[field] = self.worksheet[field].value

            if col_letter == fields_to:  # Finish work for Row
                self.data_for_save.append({cur_row: fields_data})
                fields_data = {}
                cur_row += 1

                if rows_except:  # Excpect
                    while cur_row in rows_except:
                        cur_row += 1

                count = stop_count

                if cur_row == (rows_to + 1):
                    is_loop = False
            else:
                count += 1
            sleep(0.01)

    def __save_one_object__(self, index: int) -> dict:
        """
        Save 1 complete line to a temporary variable
        with data and exceptions \\
        **Launch via the «load» command (inside)**

        PARAMETERS:
        - **index** : Index for already extracted data from EXCEL file

        RETURN: {"except": list, "data": dictionary}
        """

        def get_field_value(
            dct: self.data_for_save
        ) -> str | int | float | None:
            """
            Gets data from 1 field

            PARAMETERS:
            - **dct** : Data from the extracted EXCEL file

            RETURN: NOTHING or VALUE
            """

            try:
                key = list(dct[index].keys())[0]
                data_key = list(dct[index][key].keys())[0]

                field = dct[index][key][data_key]
                del dct[index][key][data_key]

                field = {
                    "message": True,
                    "result": field if field != "#DIV/0!" else 0
                }
                if "".join(filter(str.isalpha, data_key)) in fields_percentage:
                    field["result"] *= 100
                return field
            except IndexError or KeyError or TypeError or AttributeError or Exception:
                field = {"message": False}
                return field

        fields_percentage: list[str] | None = [
            self.data_schema["fields_percentage"]
            if self.data_schema["fields_percentage"] else []
        ][0]
        save: dict = self.data_schema["save"]

        result: dict = {"except": [], "data": {}}
        result_time_for_except: dict = {}
        is_except: list | None = []

        for save_type, save_key in save.items():
            if save_key:  # List with keys
                result_time_for_data = {}
                for sk in save_key:
                    field = get_field_value(self.data_for_save)
                    if field["message"]:
                        result_time_for_data[sk] = field["result"]
                    else:
                        is_except.append(sk)
                        result_time_for_except[save_type] = is_except
                    sleep(0.01)
                result["data"][save_type] = result_time_for_data
                sleep(0.01)
            else:  # None
                field = get_field_value(self.data_for_save)
                if field["message"]:
                    result["data"][save_type] = field["result"]
                else:
                    is_except = None
                    result_time_for_except[save_type] = is_except
            sleep(0.01)

        # Keys left from JSON
        if len(result_time_for_except) >= 1:
            result["except"].append(result_time_for_except)
            result_time_for_except, is_except = {}, []

        return result

    def __save_from_load__(self) -> None:
        """
        Starts getting data from each record and then saves it
        to a JSON file \\
        **Launch via the «load» command**
        """

        result: list[dict] = [{"except": []}]
        len_data_for_save: int = len(self.data_for_save)
        for index in range(len_data_for_save):  # For rows
            # line = list(self.data_for_save[index].keys())[0]
            # print(f"Line No.{ line } has begun processing")
            box = self.__save_one_object__(index)
            result.append(box["data"])
            result[0]["except"] = box["except"]
            sleep(0.01)

        if len(self.data_for_save) >= 1:  # Keys left from Excel
            total: dict = {}
            str_key: str = ""

            for data in self.data_for_save:
                for key, value in data.items():
                    for k, _ in value.items():
                        str_key += " " + "".join(filter(str.isalpha, k))
                        sleep(0.01)
                    if len(str_key) >= 1:
                        total[key] = str_key.strip()
                        str_key = ""
                    sleep(0.01)
                sleep(0.01)
            if len(total) >= 1:
                result[0]["except"].append(total)

        # Except (if has)
        filename: str = f"{ self.result_filename_for_json }"
        filename += f"-except-{ self.sheet }.json"
        if len(result[0]["except"]) >= 1:
            with open(filename, "w") as file:
                dump(result[0], file, indent=2, ensure_ascii=False)

            # text = "\nDuring processing, it was found that the number of "
            # text += "Excel fields does not match those required for JSON. \n"
            # text += f"More details are in the file '{ filename }'\n"
            # print(text)
        else:
            if path.exists(filename):
                remove(filename)
                # text = "\nThe number of fields in Excel and JSON are equal. "
                # text += f"The old file '{ filename }' has been deleted.\n"
                # print(text)

        # Save
        filename: str = f"{ self.result_filename_for_json }"
        filename += f"-{ self.sheet }.json"
        with open(filename, "w") as file:
            dump(result[1:], file, indent=2, ensure_ascii=False)
        # print(f"The file is saved. File name '{ filename }'")

    def __exit__(self) -> None:
        """
        Closing an EXCEL file \\
        **Launch via the «close» command**
        """

        if self.workbook:
            self.workbook.close()
